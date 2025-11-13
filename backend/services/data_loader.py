from __future__ import annotations

from pathlib import Path
from functools import lru_cache
from typing import List, Dict, Any, Optional

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[2]
BACKEND_DIR = Path(__file__).resolve().parents[1]
DATA_DIR = BACKEND_DIR / "data"

DATA_DIR.mkdir(exist_ok=True)


def _excel_path(filename: str) -> Path:
    # Mencari file Excel di root repo
    p = ROOT / filename
    if not p.exists():
        # Coba juga di dalam folder backend/data jika sudah diekspor
        q = DATA_DIR / filename
        if q.exists():
            return q
    return p


def _sheet_to_records(ws) -> List[Dict[str, Any]]:
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    out: List[Dict[str, Any]] = []
    for r in rows[1:]:
        item = {}
        for i, h in enumerate(headers):
            # openpyxl already returns Python primitives
            item[h or f"col_{i}"] = r[i] if i < len(r) else None
        out.append(item)
    return out


@lru_cache(maxsize=16)
def load_excel_as_records(filename: str, sheet_name: Optional[str] = None) -> List[Dict[str, Any]]:
    path = _excel_path(filename)
    if not path.exists():
        raise FileNotFoundError(f"File tidak ditemukan: {path}")

    wb = load_workbook(filename=str(path), read_only=True, data_only=True)
    if sheet_name is None:
        ws = wb[wb.sheetnames[0]]
        return _sheet_to_records(ws)
    else:
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            return _sheet_to_records(ws)
        # jika nama sheet tidak ada, gabungkan semua
        records: List[Dict[str, Any]] = []
        for name in wb.sheetnames:
            records.extend(_sheet_to_records(wb[name]))
        return records


def export_excel_to_json(filename: str, out_json: Path, sheet_name: Optional[str] = None) -> Path:
    records = load_excel_as_records(filename, sheet_name)
    out_json.parent.mkdir(parents=True, exist_ok=True)
    import json

    with out_json.open("w", encoding="utf-8") as f:
        json.dump(records, f, ensure_ascii=False, indent=2)
    return out_json
